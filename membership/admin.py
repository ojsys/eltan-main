from django import forms
from django.contrib import admin
#from  django_summernote.admin import SummernoteModelAdmin
from django.utils.html import format_html
from django.utils import timezone
from django.db import OperationalError as DBOperationalError
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from .email_utils import send_bulk_async
from django.db import models
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from .models import Conference, ConferenceRegistration, EltanConference, EltanConferenceRegistration, ConferenceDocument, MemberProfile, MembershipType, Subscription, Sigs, SigsRegistration, Events, News, Resource, Download, ELTANYearSetting, Newsletter, ConferenceSpeaker, ConferenceSchedule, ConferenceSponsor, ConferenceLocMember, ExcoMember, SponsorshipPackage, ConferenceAccommodation, CertificateSignatory, normalize_eltan_year



@admin.register(ELTANYearSetting)
class ELTANYearSettingAdmin(admin.ModelAdmin):
    """Manage the ELTAN years members can subscribe to.

    Creating a year is meant to be one step: type the label (e.g. 2026-2027) and
    save. The dates fill themselves in, and the year appears in the member
    subscription form immediately.
    """

    list_display = ('eltan_year', 'current_badge', 'is_selectable', 'start_date', 'end_date', 'subscription_count')
    list_filter = ('is_active', 'is_selectable')
    list_editable = ('is_selectable',)
    search_fields = ('eltan_year',)
    ordering = ('-eltan_year',)
    readonly_fields = ('created_at', 'updated_at')
    actions = ['make_current', 'add_next_year', 'make_selectable', 'make_unselectable']

    fieldsets = (
        (None, {
            'fields': ('eltan_year',),
            'description': (
                'Type the year as <strong>2026-2027</strong> (two consecutive years). '
                'Save and it is immediately available to members on the subscription form.'
            ),
        }),
        ('Dates', {
            'fields': ('start_date', 'end_date'),
            'description': 'Leave both blank to use the standard 1 September – 31 August span.',
            'classes': ('collapse',),
        }),
        ('Availability', {
            'fields': ('is_active', 'is_selectable'),
        }),
        ('Record', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )

    def current_badge(self, obj):
        if obj.is_active:
            return format_html(
                '<span style="background:#f0fdf4; color:#16a34a; border:1px solid #16a34a; '
                'padding:3px 10px; border-radius:12px; font-size:11px; font-weight:600;">CURRENT</span>'
            )
        return format_html('<span style="color:#9ca3af; font-size:12px;">—</span>')
    current_badge.short_description = 'Current year'

    def subscription_count(self, obj):
        return Subscription.objects.filter(eltan_year=obj.eltan_year).count()
    subscription_count.short_description = 'Subscriptions'

    def make_current(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(request, 'Select exactly one year to make current.', messages.ERROR)
            return
        year = queryset.first()
        year.is_active = True
        year.is_selectable = True
        year.save()  # model save() stands the other years down
        self.message_user(request, f'{year.eltan_year} is now the current ELTAN year.', messages.SUCCESS)
    make_current.short_description = 'Set as the current ELTAN year'

    def add_next_year(self, request, queryset):
        """Create the year after the newest one, so opening a new ELTAN year is
        a single click rather than a code change."""
        newest = ELTANYearSetting.objects.order_by('-eltan_year').first()
        if not newest or newest.start_year is None:
            self.message_user(request, 'No valid existing year to follow on from.', messages.ERROR)
            return
        next_start = newest.start_year + 1
        label = f'{next_start}-{next_start + 1}'
        year, created = ELTANYearSetting.objects.get_or_create(
            eltan_year=label, defaults={'is_active': False, 'is_selectable': True},
        )
        if created:
            self.message_user(request, f'ELTAN year {label} created and open for selection.', messages.SUCCESS)
        else:
            self.message_user(request, f'ELTAN year {label} already exists.', messages.INFO)
    add_next_year.short_description = 'Create the next ELTAN year'

    def make_selectable(self, request, queryset):
        updated = queryset.update(is_selectable=True)
        self.message_user(request, f'{updated} year(s) are now selectable by members.', messages.SUCCESS)
    make_selectable.short_description = 'Show on the subscription form'

    def make_unselectable(self, request, queryset):
        updated = queryset.filter(is_active=False).update(is_selectable=False)
        skipped = queryset.filter(is_active=True).count()
        self.message_user(
            request,
            f'{updated} year(s) hidden from the subscription form.'
            + (' The current year cannot be hidden.' if skipped else ''),
            messages.SUCCESS if updated else messages.WARNING,
        )
    make_unselectable.short_description = 'Hide from the subscription form'


# class NewsAdmin(SummernoteModelAdmin):
#     summernote_fields = ('content',)
class MemberProfileAdmin(admin.ModelAdmin):
    list_display = ['user', 'gender', 'phone_number', 'state', 'city']
    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Member Profiles"

        # Define headers
        headers = [
            'Full Name', 'Email', 'Gender', 'Phone Number', 'Address',
            'City', 'State', 'Country', 'Date of Birth', 'Created At'
        ]
        ws.append(headers)

        # Add data
        for obj in queryset:
            row = [
                f"{obj.user.first_name} {obj.user.last_name}",
                obj.user.email,
                obj.gender,
                obj.phone_number,
                obj.address,
                obj.city,
                obj.state,
                obj.country,
                obj.date_of_birth.strftime('%Y-%m-%d') if obj.date_of_birth else '',
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            ws.append(row)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=member_profiles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response

    export_to_excel.short_description = "Export selected members to Excel"



class ELTANYearChoiceField(forms.ChoiceField):
    """A year picker that treats '2025/2026' and '2025-2026' as the same choice.

    Normalising in ``to_python`` rather than ``clean_<field>`` matters: a
    ChoiceField checks the raw value against its choices first, so a legacy
    slashed year would be rejected as "not one of the available choices" before
    any later hook got the chance to canonicalise it.
    """

    def to_python(self, value):
        return normalize_eltan_year(super().to_python(value))


class SubscriptionAdminForm(forms.ModelForm):
    """Admin form for a subscription.

    ``eltan_year`` is a plain CharField on the model (the valid years live in
    ELTANYearSetting rather than in hardcoded choices), which left the admin with
    a free-text box — one typo there and the subscription belonged to a year that
    does not exist. Offer the configured years as a dropdown instead.
    """

    class Meta:
        model = Subscription
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        years = list(ELTANYearSetting.objects.order_by('-eltan_year').values_list('eltan_year', flat=True))

        # Keep whatever this row already has, even if the year was since removed
        # or is a legacy label — editing an old subscription must not be blocked
        # by an unrelated field, and saving must not silently change its year.
        current = normalize_eltan_year(self.instance.eltan_year if self.instance else '')
        if current and current not in years:
            years.insert(0, current)

        self.fields['eltan_year'] = ELTANYearChoiceField(
            label='ELTAN Year',
            required=False,
            choices=[('', '--- Select ELTAN year ---')] + [(year, year) for year in years],
            initial=current or ELTANYearSetting.current_label(),
            help_text='Years come from ELTAN Years — add one there to see it here.',
        )


class SubscriptionAdmin(admin.ModelAdmin):
    form = SubscriptionAdminForm
    list_display = [
        'user', 'get_eltan_number', 'membership_type', 'eltan_year', 'start_date', 'end_date',
        'payment_status', 'amount_paid', 'cert_status_badge', 'receipt_status',
        'qualification_cert_link', 'payment_proof_thumbnail',
    ]
    actions = [
        'recalculate_eltan_dates', 'send_subscription_receipts', 'approve_certificate',
        'reject_certificate', 'export_to_excel',
    ]

    search_fields = ['user__first_name', 'user__last_name', 'user__email', 'user__eltan_number',
                     'membership_type', 'payment_status', 'state_chapter']

    list_filter = ['payment_status', 'certificate_status', 'membership_type', 'eltan_year', 'state_chapter']

    readonly_fields = ('payment_proof_full', 'qualification_cert_preview')

    def get_eltan_number(self, obj):
        return obj.user.eltan_number
    get_eltan_number.short_description = 'ELTAN Number'

    def save_model(self, request, obj, form, change):
        """Email the confirmed receipt when staff mark a payment as paid.

        A member who paid by bank transfer gets their 'we have your subscription'
        receipt when they submit it; the confirmation is owed at the moment staff
        verify the transfer, which happens here. Only on the transition, so
        re-saving an already-paid subscription does not send it twice.
        """
        became_paid = obj.payment_status == 'paid' and 'payment_status' in form.changed_data
        super().save_model(request, obj, form, change)

        if not became_paid:
            return

        # Imported here rather than at module level: views imports models, and
        # models is what loads this admin, so a top-level import would be circular.
        from .views import send_subscription_receipt

        ok, error = send_subscription_receipt(obj)
        if ok:
            self.message_user(request, f'Receipt emailed to {obj.user.email}.', messages.SUCCESS)
        else:
            self.message_user(
                request,
                f'Subscription saved, but the receipt could not be emailed: {error}',
                messages.WARNING,
            )

    def amount_paid(self, obj):
        return obj.payment_amount
    amount_paid.short_description = 'Amount Paid'

    def cert_status_badge(self, obj):
        colors = {
            'pending':  ('#f59e0b', '#fffbeb', 'PENDING'),
            'approved': ('#16a34a', '#f0fdf4', 'APPROVED'),
            'rejected': ('#dc2626', '#fef2f2', 'REJECTED'),
        }
        color, bg, label = colors.get(obj.certificate_status, ('#6b7280', '#f9fafb', obj.certificate_status.upper()))
        return format_html(
            '<span style="background:{}; color:{}; padding:3px 10px; border-radius:12px; '
            'font-size:11px; font-weight:600; border:1px solid {};">{}</span>',
            bg, color, color, label,
        )
    cert_status_badge.short_description = 'Cert Status'

    def receipt_status(self, obj):
        """Whether the member actually received their receipt."""
        if obj.receipt_sent_at:
            return format_html(
                '<span style="color:#16a34a; font-size:12px;" title="Sent {}">&#10003; Sent</span>',
                obj.receipt_sent_at.strftime('%Y-%m-%d %H:%M'),
            )
        if obj.receipt_error:
            return format_html(
                '<span style="color:#dc2626; font-size:12px;" title="{}">&#10007; Failed</span>',
                obj.receipt_error[:300],
            )
        return format_html('<span style="color:#9ca3af; font-size:12px;">Not sent</span>')
    receipt_status.short_description = 'Receipt'

    def qualification_cert_link(self, obj):
        if obj.qualification_certificate:
            return format_html(
                '<a href="{}" target="_blank" style="font-size:12px;">View Certificate</a>',
                obj.qualification_certificate.url,
            )
        return format_html('<span style="color:#9ca3af; font-size:12px;">None uploaded</span>')
    qualification_cert_link.short_description = 'Qualification Cert'

    def payment_proof_thumbnail(self, obj):
        if obj.payment_proof:
            return format_html(
                '<a href="#" onclick="showPaymentProof(\'{}\', event)" class="payment-proof-link">'
                '<img src="{}" style="max-height: 50px; border-radius: 4px; border: 1px solid #ddd;"/>'
                '</a>',
                obj.payment_proof.url,
                obj.payment_proof.url,
            )
        return "No proof uploaded"
    payment_proof_thumbnail.short_description = 'Payment Proof'

    def payment_proof_full(self, obj):
        if obj.payment_proof:
            return format_html(
                '<div style="max-width: 100%; margin-top: 10px;">'
                '<img src="{}" style="max-width: 100%; max-height: 600px; border-radius: 8px;"/>'
                '</div>',
                obj.payment_proof.url,
            )
        return "No proof uploaded"
    payment_proof_full.short_description = 'Payment Proof (Full)'

    def qualification_cert_preview(self, obj):
        if obj.qualification_certificate:
            url = obj.qualification_certificate.url
            name = obj.qualification_certificate.name.split('/')[-1]
            return format_html(
                '<a href="{}" target="_blank" style="display:inline-block; padding:8px 16px; '
                'background:#1d4ed8; color:white; border-radius:6px; text-decoration:none; '
                'font-size:13px;">Open: {}</a>',
                url, name,
            )
        return format_html('<span style="color:#9ca3af;">No certificate uploaded</span>')
    qualification_cert_preview.short_description = 'Qualification Certificate'

    # --- Actions ---

    def recalculate_eltan_dates(self, request, queryset):
        """Reset end_date to what the membership rules say it should be: the end
        of the ELTAN year for a first membership, 365 days from the start date
        for a renewal."""
        updated = 0
        for subscription in queryset:
            dates = subscription.calculate_eltan_dates()
            if not dates['end_date']:
                continue
            subscription.end_date = dates['end_date']
            subscription.eltan_year = dates['eltan_year']
            subscription.save(update_fields=['end_date', 'eltan_year'])
            updated += 1
        self.message_user(
            request,
            f'{updated} subscription(s) recalculated: first memberships end with the '
            f'ELTAN year, renewals run 365 days from their start date.',
            messages.SUCCESS,
        )
    recalculate_eltan_dates.short_description = 'Recalculate end date (ELTAN year end, or 365 days for renewals)'

    def send_subscription_receipts(self, request, queryset):
        """(Re)send the subscription receipt to the selected members.

        For a member who never got theirs — a mail outage, a bad address since
        corrected — or when they simply ask for another copy.
        """
        # Imported here rather than at module level: views imports models, and
        # models is what loads this admin, so a top-level import would be circular.
        from .views import send_subscription_receipt

        sent = 0
        failures = []
        for subscription in queryset:
            ok, error = send_subscription_receipt(subscription)
            if ok:
                sent += 1
            else:
                failures.append(f'{subscription.user.email}: {error}')

        if sent:
            self.message_user(request, f'Receipt emailed to {sent} member(s).', messages.SUCCESS)
        if failures:
            self.message_user(
                request,
                f'{len(failures)} receipt(s) failed — ' + '; '.join(failures[:5]),
                messages.ERROR,
            )
    send_subscription_receipts.short_description = 'Email subscription receipt to selected members'

    def approve_certificate(self, request, queryset):
        approved = queryset.filter(certificate_status__in=['pending', 'rejected'])
        emails = []
        count = 0
        for subscription in approved:
            subscription.certificate_status = 'approved'
            subscription.save(update_fields=['certificate_status'])
            email = self._build_cert_verification_email(subscription, approved=True)
            if email:
                emails.append(email)
            count += 1
        send_bulk_async(emails)
        self.message_user(
            request,
            f'{count} subscription(s) approved — {len(emails)} notification email(s) '
            f'are being sent in the background.',
            messages.SUCCESS,
        )
    approve_certificate.short_description = 'Approve certificate & activate subscription'

    def reject_certificate(self, request, queryset):
        rejected = queryset.filter(certificate_status__in=['pending', 'approved'])
        emails = []
        count = 0
        for subscription in rejected:
            subscription.certificate_status = 'rejected'
            subscription.save(update_fields=['certificate_status'])
            email = self._build_cert_verification_email(subscription, approved=False)
            if email:
                emails.append(email)
            count += 1
        send_bulk_async(emails)
        self.message_user(
            request,
            f'{count} subscription(s) rejected — {len(emails)} notification email(s) '
            f'are being sent in the background.',
            messages.WARNING,
        )
    reject_certificate.short_description = 'Reject certificate & block subscription'

    def _build_cert_verification_email(self, subscription, approved: bool):
        """Build (but do not send) the cert verification email. Returns an
        EmailMessage, or None when the member has no email address. Sending is
        done in bulk over a single connection by send_bulk_async()."""
        user = subscription.user
        if not user or not user.email:
            return None
        name = f"{user.first_name} {user.last_name}".strip() or user.email
        site_name = getattr(settings, 'SITE_NAME', 'ELTAN')
        site_url = getattr(settings, 'SITE_URL', 'https://eltanigeria.org')

        if approved:
            subject = f"Your ELTAN Qualification Certificate Has Been Verified"
            message = (
                f"Dear {name},\n\n"
                f"We are pleased to inform you that your qualification certificate has been reviewed and approved.\n\n"
                f"Your MEMBERSHIP STATUS for {subscription.eltan_year} is now fully active "
                f"and you can download your ELTAN membership certificate from your dashboard.\n\n"
                f"Visit your dashboard: {site_url}/certificates/\n\n"
                f"Thank you for being a valued member of {site_name}.\n\n"
                f"Best regards,\n{site_name} Team"
            )
        else:
            subject = f"ELTAN Qualification Certificate — Action Required"
            message = (
                f"Dear {name},\n\n"
                f"We have reviewed your qualification certificate submitted for your "
                f"MEMBERSHIP STATUS for ({subscription.eltan_year})"
                f"and unfortunately it could not be approved at this time.\n\n"
                f"Please contact our support team for further assistance or to resubmit the correct documentation.\n\n"
                f"Best regards,\n{site_name} Team"
            )

        return EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[user.email],
        )

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Subscriptions"
        headers = [
            'Member Name', 'Email', 'Eltan Number', 'Membership Type', 'Start Date',
            'End Date', 'Payment Status', 'Certificate Status', 'Payment Amount', 'State Chapter',
        ]
        ws.append(headers)
        for obj in queryset:
            row = [
                f"{obj.user.first_name} {obj.user.last_name}",
                obj.user.email,
                obj.user.eltan_number,
                obj.membership_type,
                obj.start_date.strftime('%Y-%m-%d'),
                obj.end_date.strftime('%Y-%m-%d') if obj.end_date else '',
                obj.payment_status,
                obj.certificate_status,
                str(obj.payment_amount) if obj.payment_amount else '',
                obj.state_chapter,
            ]
            ws.append(row)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=subscriptions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    export_to_excel.short_description = "Export selected subscriptions to Excel"

    class Media:
        css = {
            'all': ('admin/css/payment_proof_modal.css',)
        }
        js = ('admin/js/payment_proof_modal.js',)



class ConferenceLocMemberInline(admin.StackedInline):
    model = ConferenceLocMember
    extra = 1
    fields = ('name', 'role', 'organization', 'email', 'phone', 'image', 'order')
    ordering = ('order', 'name')
    verbose_name = 'LOC Member'
    verbose_name_plural = 'Local Organizing Committee (LOC) Members'


class ConferenceSpeakerInline(admin.StackedInline):
    model = ConferenceSpeaker
    extra = 0
    fields = ('name', 'title', 'image', 'presentation_title', 'order')
    ordering = ('order', 'name')
    verbose_name = 'Speaker'
    verbose_name_plural = 'Speakers'


class ConferenceScheduleInline(admin.TabularInline):
    model = ConferenceSchedule
    extra = 0
    fields = ('date', 'start_time', 'end_time', 'session_title', 'location', 'speaker')


class ConferenceDocumentInline(admin.TabularInline):
    model = ConferenceDocument
    extra = 0
    fields = ('title', 'document', 'is_public')


class SponsorshipPackageInline(admin.StackedInline):
    model = SponsorshipPackage
    extra = 0
    fields = ('tier', 'tier_label', 'price_range', 'benefits', 'is_featured', 'cta_label', 'order')
    verbose_name = 'Sponsorship Package'
    verbose_name_plural = 'Sponsorship Packages'


class ConferenceAccommodationInline(admin.StackedInline):
    model = ConferenceAccommodation
    extra = 0
    fields = (
        'name', 'address', 'distance_from_venue', 'price_range', 'room_types',
        'contact_phone', 'contact_email', 'website', 'booking_deadline',
        'notes', 'is_recommended', 'image', 'order',
    )
    verbose_name = 'Accommodation'
    verbose_name_plural = 'Accommodation Options'


@admin.register(SponsorshipPackage)
class SponsorshipPackageAdmin(admin.ModelAdmin):
    list_display = ('tier_label', 'conference', 'price_range', 'is_featured', 'order')
    list_filter = ('conference', 'tier')
    ordering = ('conference', 'order')


@admin.register(ConferenceAccommodation)
class ConferenceAccommodationAdmin(admin.ModelAdmin):
    list_display = ('name', 'conference', 'distance_from_venue', 'price_range', 'is_recommended', 'order')
    list_filter = ('conference', 'is_recommended')
    search_fields = ('name', 'address')
    ordering = ('conference', 'order', 'name')


@admin.register(EltanConference)
class EltanConferenceAdmin(admin.ModelAdmin):
    list_display = ('title', 'start_date', 'end_date', 'registration_status', 'is_active', 'member_fee', 'non_member_fee')
    readonly_fields = ('is_early_bird_active',)
    list_filter = ('is_active', 'start_date')
    search_fields = ('title', 'theme', 'description')
    actions = None
    inlines = [ConferenceLocMemberInline, ConferenceSpeakerInline, ConferenceScheduleInline, ConferenceDocumentInline, SponsorshipPackageInline, ConferenceAccommodationInline]
    fieldsets = (
        ('Conference Info', {
            'fields': ('title', 'theme', 'description', 'image', 'is_active')
        }),
        ('Dates & Venue', {
            'fields': ('start_date', 'end_date', 'venue')
        }),
        ('Registration', {
            'fields': ('registration_start', 'registration_end', 'early_bird_end', 'is_early_bird_active',
                       'abstract_form_link')
        }),
        ('Fees', {
            'fields': ('member_fee', 'member_early_bird_fee', 'non_member_fee', 'non_member_early_bird_fee',
                       'international_delegate_fee', 'virtual_attendee_fee')
        }),
        ('Payment Links', {
            'fields': ('member_payment_link', 'non_member_payment_link'),
            'classes': ('collapse',),
        }),
        ('Contact', {
            'fields': ('contact_name', 'contact_email', 'contact_phone'),
        }),
        ('Content Tabs', {
            'fields': ('sub_themes', 'cfp_guidelines', 'sponsor_packages'),
            'description': 'These fields populate the Sub-Themes, Call for Papers, and Sponsors tabs on the conference portal.',
        }),
    )
    
    def get_queryset(self, request):
        """Avoid selecting newly added columns if migrations haven't been applied yet."""
        qs = super().get_queryset(request)
        try:
            # Force a lightweight evaluation to detect schema errors early
            list(qs.values_list('pk', flat=True)[:1])
            return qs.only(
                'id', 'title', 'theme', 'image', 'start_date', 'end_date', 'venue',
                'registration_start', 'registration_end', 'early_bird_end', 'is_active',
                'member_fee', 'member_early_bird_fee', 'non_member_fee', 'non_member_early_bird_fee',
                'international_delegate_fee', 'sponsor_packages'
            )
        except DBOperationalError:
            messages.error(request, 'Conference list is limited until database migrations are applied. Please run migrations to enable new fields.')
            return self.model.objects.none()

    def changelist_view(self, request, extra_context=None):
        try:
            return super().changelist_view(request, extra_context=extra_context)
        except DBOperationalError:
            messages.error(request, 'Conference list is unavailable until database migrations are applied for new fields.')
            from django.shortcuts import redirect
            return redirect('admin:index')

    def get_search_results(self, request, queryset, search_term):
        try:
            return super().get_search_results(request, queryset, search_term)
        except DBOperationalError:
            messages.error(request, 'Search disabled until database migrations are applied for conference fields.')
            return self.model.objects.none(), False
    
    def registration_status(self, obj):
        if obj.is_open_for_registration:
            return 'Open'
        elif timezone.now().date() < obj.registration_start:
            return 'Not Started'
        else:
            return 'Closed'
    registration_status.short_description = 'Registration Status'
    
    def save_model(self, request, obj, form, change):
        if not obj.member_fee or not obj.non_member_fee:
            raise ValidationError("All fee fields must be set")
        super().save_model(request, obj, form, change)


@admin.register(ConferenceLocMember)
class ConferenceLocMemberAdmin(admin.ModelAdmin):
    list_display = ('name', 'role', 'conference', 'order')
    list_filter = ('conference',)
    search_fields = ('name', 'role', 'organization', 'email', 'phone')
    ordering = ('conference', 'order', 'name')


class UserTypeFilter(admin.SimpleListFilter):
    title = 'User Type'  # Display name for the filter
    parameter_name = 'user_type'  # Query parameter name

    def lookups(self, request, model_admin):
        """
        Define filter options.
        """
        return (
            ('user', 'Authenticated User'),
            ('non_user', 'Non-User'),
        )

    def queryset(self, request, queryset):
        """
        Filter the queryset based on the selected option.
        """
        if self.value() == 'user':
            return queryset.filter(user__isnull=False)
        if self.value() == 'non_user':
            return queryset.filter(user__isnull=True)
        return queryset


class ReceiptStatusFilter(admin.SimpleListFilter):
    """Find the people who paid but never received their ticket."""

    title = 'Ticket email'
    parameter_name = 'receipt'

    def lookups(self, request, model_admin):
        return (
            ('missing', 'Paid — ticket NOT sent'),
            ('sent', 'Paid — ticket sent'),
            ('failed', 'Last send failed'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'missing':
            return queryset.filter(payment_status='completed', receipt_sent_at__isnull=True)
        if self.value() == 'sent':
            return queryset.filter(payment_status='completed', receipt_sent_at__isnull=False)
        if self.value() == 'failed':
            return queryset.exclude(receipt_error='')
        return queryset


@admin.register(EltanConferenceRegistration)
class EltanConferenceRegistrationAdmin(admin.ModelAdmin):
    list_display = ('display_user', 'conference', 'registration_type', 'payment_status', 'ticket_id', 'amount_paid',
                    'receipt_status', 'is_presenting', 'registered_at', 'phone')
    list_filter = (UserTypeFilter, ReceiptStatusFilter, 'conference', 'registration_type', 'payment_status', 'is_presenting')

    def receipt_status(self, obj):
        """Whether the ticket/receipt email actually reached the attendee."""
        if obj.payment_status != 'completed':
            return format_html('<span style="color:#9ca3af; font-size:12px;">—</span>')
        if obj.receipt_sent_at:
            return format_html(
                '<span style="background:#f0fdf4; color:#16a34a; border:1px solid #16a34a; padding:3px 10px; '
                'border-radius:12px; font-size:11px; font-weight:600;" title="{}">SENT</span>',
                obj.receipt_sent_at.strftime('%d %b %Y %H:%M'),
            )
        return format_html(
            '<span style="background:#fef2f2; color:#dc2626; border:1px solid #dc2626; padding:3px 10px; '
            'border-radius:12px; font-size:11px; font-weight:600;" title="{}">NOT SENT</span>',
            obj.receipt_error or 'Never attempted',
        )
    receipt_status.short_description = 'Ticket email'

    def display_user(self, obj):
        """
        Custom method to display user or non-user details.
        """
        if obj.user:
            return obj.user  # Display the authenticated user
        else:
            # Display non-user details
            return f"{obj.first_name} {obj.last_name} ({obj.email})"
    
    display_user.short_description = 'Member / Non-Member'
    
    search_fields = ('user__email', 'user__first_name', 'user__last_name',
                    'paper_title', 'email', 'first_name', 'last_name', 'ticket_id')  # Added non-user fields
    readonly_fields = ('registered_at', 'ticket_id', 'payment_verified_at', 'verified_by',
                       'receipt_sent_at', 'receipt_error')

    actions = ['verify_payments', 'resend_tickets', 'export_registrations']

    def verify_payments(self, request, queryset):
        from .views import send_registration_receipt
        confirmed = 0
        skipped = 0
        failed_emails = []
        for registration in queryset:
            if registration.payment_status == 'completed':
                skipped += 1
                continue
            registration.mark_completed(verified_by=request.user)
            ok, error = send_registration_receipt(registration)
            if not ok:
                failed_emails.append(f"{registration.contact_email or f'#{registration.pk}'} ({error})")
            confirmed += 1

        self.message_user(
            request,
            f"{confirmed} registration(s) confirmed and ticketed; {skipped} already completed.",
            messages.SUCCESS,
        )
        if failed_emails:
            # Never let a silent mail failure pass for success — these people paid
            # and are waiting for a ticket that did not go out.
            self.message_user(
                request,
                f"{len(failed_emails)} ticket email(s) could NOT be sent: "
                + '; '.join(failed_emails[:5])
                + ('…' if len(failed_emails) > 5 else '')
                + " Fix the mail settings, then use 'Resend ticket email'.",
                messages.ERROR,
            )
    verify_payments.short_description = "Verify & confirm selected payments (issue ticket + email)"

    def resend_tickets(self, request, queryset):
        """Re-send the ticket/receipt for confirmed registrations — the catch-up
        for anyone who paid while mail was misconfigured."""
        from .views import send_registration_receipt
        sent = 0
        failures = []
        not_confirmed = 0
        for registration in queryset:
            if registration.payment_status != 'completed':
                not_confirmed += 1
                continue
            ok, error = send_registration_receipt(registration)
            if ok:
                sent += 1
            else:
                failures.append(f"{registration.contact_email or f'#{registration.pk}'} ({error})")

        self.message_user(
            request,
            f"{sent} ticket email(s) sent."
            + (f" {not_confirmed} skipped (payment not confirmed)." if not_confirmed else ''),
            messages.SUCCESS if sent else messages.WARNING,
        )
        if failures:
            self.message_user(
                request,
                f"{len(failures)} failed: " + '; '.join(failures[:5]) + ('…' if len(failures) > 5 else ''),
                messages.ERROR,
            )
    resend_tickets.short_description = "Resend ticket / receipt email"

    def export_registrations(self, request, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="conference_registrations.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Ticket ID', 'Name', 'Email', 'Registration Type', 'Payment Status',
            'Amount Paid', 'Is Presenting', 'Paper Title', 'Phone',
            'Registration Date'
        ])
        
        for registration in queryset:
            # Get name and email based on whether it's a member or non-member
            if registration.user:
                name = f"{registration.user.first_name} {registration.user.last_name}"
                email = registration.user.email
            else:
                name = f"{registration.first_name} {registration.last_name}"
                email = registration.email
            
            writer.writerow([
                registration.ticket_id or '',
                name,
                email,
                registration.registration_type,
                registration.payment_status,
                registration.amount_paid,
                'Yes' if registration.is_presenting else 'No',
                registration.paper_title or '',
                registration.phone or '',
                registration.registered_at.strftime('%Y-%m-%d %H:%M:%S') if registration.registered_at else ''
            ])
        
        return response
    export_registrations.short_description = "Export selected registrations to CSV"

@admin.register(ConferenceDocument)
class ConferenceDocumentAdmin(admin.ModelAdmin):
    list_display = ('title', 'conference', 'is_public', 'uploaded_at')
    list_filter = ('conference', 'is_public')
    search_fields = ('title',)


@admin.register(ConferenceSpeaker)
class ConferenceSpeakerAdmin(admin.ModelAdmin):
    list_display = ('name', 'title', 'conference', 'presentation_title')
    list_filter = ('conference',)
    search_fields = ('name', 'title', 'presentation_title')
    ordering = ('conference', 'order', 'name')


@admin.register(ConferenceSchedule)
class ConferenceScheduleAdmin(admin.ModelAdmin):
    list_display = ('session_title', 'conference', 'date', 'start_time', 'end_time', 'location', 'speaker')
    list_filter = ('conference', 'date')
    search_fields = ('session_title', 'location', 'speaker__name')
    ordering = ('date', 'start_time')


@admin.register(ConferenceSponsor)
class ConferenceSponsorAdmin(admin.ModelAdmin):
    list_display = ('company_name', 'conference', 'level', 'contact_name', 'is_approved')
    list_filter = ('conference', 'level', 'is_approved')
    search_fields = ('company_name', 'contact_name', 'contact_email')
    list_editable = ('is_approved',)


@admin.register(ExcoMember)
class ExcoMemberAdmin(admin.ModelAdmin):
    list_display = (
        'photo_preview',
        'full_name',
        'position',
        'institution',
        'tenure_period',
        'status_badge',
        'order'
    )

    list_filter = (
        'is_active',
        'is_archived',
        'position',
        'start_date',
    )

    search_fields = (
        'full_name',
        'position',
        'institution',
        'email',
        'user__first_name',
        'user__last_name',
        'user__email',
    )

    list_editable = ('order',)

    ordering = ('is_archived', '-is_active', 'order', 'position')

    actions = ['archive_members', 'unarchive_members', 'mark_as_active', 'export_to_excel']

    autocomplete_fields = ['user']

    readonly_fields = ('created_at', 'updated_at', 'photo_preview_large')

    fieldsets = (
        ('Member Information', {
            'fields': (
                'user',
                'full_name',
                'position',
                'institution',
            )
        }),
        ('Contact Details', {
            'fields': (
                'email',
                'phone',
                'linkedin_url',
            )
        }),
        ('Photo & Bio', {
            'fields': (
                'photo',
                'photo_preview_large',
                'bio',
            )
        }),
        ('Tenure & Status', {
            'fields': (
                'start_date',
                'end_date',
                'is_active',
                'is_archived',
                'order',
            )
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def photo_preview(self, obj):
        """Small thumbnail for list view"""
        photo_url = obj.get_photo_url()
        if photo_url:
            return format_html(
                '<img src="{}" width="50" height="50" style="border-radius: 50%; object-fit: cover; border: 2px solid #ddd;"/>',
                photo_url
            )
        return format_html(
            '<div style="width: 50px; height: 50px; border-radius: 50%; background: #e0e0e0; '
            'display: flex; align-items: center; justify-content: center; color: #666; font-size: 20px;">{}</div>',
            obj.full_name[0] if obj.full_name else '?'
        )
    photo_preview.short_description = 'Photo'

    def photo_preview_large(self, obj):
        """Large preview for detail view"""
        photo_url = obj.get_photo_url()
        if photo_url:
            return format_html(
                '<img src="{}" style="max-width: 300px; max-height: 300px; border-radius: 8px; border: 1px solid #ddd;"/>',
                photo_url
            )
        return "No photo uploaded"
    photo_preview_large.short_description = 'Photo Preview'

    def tenure_period(self, obj):
        """Display tenure as a date range"""
        start = obj.start_date.strftime('%b %Y')
        if obj.end_date:
            end = obj.end_date.strftime('%b %Y')
            return f"{start} - {end}"
        return f"{start} - Present"
    tenure_period.short_description = 'Tenure'

    def status_badge(self, obj):
        """Visual status indicator"""
        if obj.is_archived:
            return format_html(
                '<span style="background: #9e9e9e; color: white; padding: 3px 10px; '
                'border-radius: 12px; font-size: 11px; font-weight: 600;">ARCHIVED</span>'
            )
        elif obj.is_active:
            return format_html(
                '<span style="background: #4caf50; color: white; padding: 3px 10px; '
                'border-radius: 12px; font-size: 11px; font-weight: 600;">CURRENT</span>'
            )
        else:
            return format_html(
                '<span style="background: #ff9800; color: white; padding: 3px 10px; '
                'border-radius: 12px; font-size: 11px; font-weight: 600;">INACTIVE</span>'
            )
    status_badge.short_description = 'Status'

    def archive_members(self, request, queryset):
        """Archive selected members (move to Past Exco)"""
        updated = queryset.update(is_archived=True, is_active=False)
        self.message_user(
            request,
            f'{updated} member(s) successfully archived and moved to Past Exco.',
            messages.SUCCESS
        )
    archive_members.short_description = "Archive selected members (move to Past Exco)"

    def unarchive_members(self, request, queryset):
        """Unarchive selected members"""
        updated = queryset.update(is_archived=False)
        self.message_user(
            request,
            f'{updated} member(s) unarchived. Remember to set is_active if they should appear in Current Exco.',
            messages.WARNING
        )
    unarchive_members.short_description = "Unarchive selected members"

    def mark_as_active(self, request, queryset):
        """Mark as active (current exco)"""
        updated = queryset.update(is_active=True, is_archived=False)
        self.message_user(
            request,
            f'{updated} member(s) marked as active in Current Exco.',
            messages.SUCCESS
        )
    mark_as_active.short_description = "Mark as active (Current Exco)"

    def export_to_excel(self, request, queryset):
        """Export to Excel following existing pattern"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Exco Members"

        headers = [
            'Full Name', 'Position', 'Institution', 'Email', 'Phone',
            'LinkedIn', 'Start Date', 'End Date', 'Status', 'Is Active', 'Is Archived'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.full_name,
                obj.get_position_display(),
                obj.institution,
                obj.get_email(),
                obj.phone,
                obj.linkedin_url,
                obj.start_date.strftime('%Y-%m-%d'),
                obj.end_date.strftime('%Y-%m-%d') if obj.end_date else 'Present',
                'Archived' if obj.is_archived else ('Active' if obj.is_active else 'Inactive'),
                'Yes' if obj.is_active else 'No',
                'Yes' if obj.is_archived else 'No',
            ]
            ws.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=exco_members_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    export_to_excel.short_description = "Export selected members to Excel"

    def get_queryset(self, request):
        """Optimize queries"""
        qs = super().get_queryset(request)
        return qs.select_related('user').prefetch_related('user__memberprofile')


@admin.register(Newsletter)
class NewsletterAdmin(admin.ModelAdmin):
    list_display = ('title', 'month_year', 'file_size', 'upload_date', 'is_published', 'is_featured')
    list_filter = ('year', 'month', 'is_published', 'is_featured')
    search_fields = ('title', 'description')
    ordering = ('-year', '-month')
    list_editable = ('is_published', 'is_featured')
    readonly_fields = ('upload_date', 'file_size')

    fieldsets = (
        ('Newsletter Information', {
            'fields': ('title', 'month', 'year', 'description')
        }),
        ('Files', {
            'fields': ('pdf_file', 'thumbnail'),
            'description': 'Upload the newsletter PDF file and an optional thumbnail image.'
        }),
        ('Publication Settings', {
            'fields': ('is_published', 'is_featured', 'upload_date', 'file_size'),
            'classes': ('collapse',)
        }),
    )

    def month_year(self, obj):
        return f"{obj.get_month_display()} {obj.year}"
    month_year.short_description = "Period"

    def save_model(self, request, obj, form, change):
        """Provide helpful feedback when saving"""
        super().save_model(request, obj, form, change)
        if not change:  # New newsletter
            self.message_user(request, f"Newsletter '{obj.title}' uploaded successfully!", messages.SUCCESS)
        else:
            self.message_user(request, f"Newsletter '{obj.title}' updated successfully!", messages.SUCCESS)



# Register your models here.
admin.site.register(MemberProfile, MemberProfileAdmin)
#admin.site.register(Conference)
#admin.site.register(ConferenceRegistration, ConferenceRegistrationAdmin)
admin.site.register(MembershipType)
admin.site.register(Subscription, SubscriptionAdmin)
# SIG Admin with inline members
class SigsRegistrationInline(admin.TabularInline):
    model = SigsRegistration
    extra = 0
    fields = ['user', 'registration_date']
    readonly_fields = ['registration_date']
    autocomplete_fields = ['user']

class SigsAdmin(admin.ModelAdmin):
    list_display = ['title', 'moderator', 'member_count', 'is_active', 'is_featured', 'order']
    list_filter = ['is_active', 'is_featured']
    search_fields = ['title', 'description']
    list_editable = ['is_active', 'is_featured', 'order']
    inlines = [SigsRegistrationInline]
    fieldsets = (
        ('Basic Information', {
            'fields': ('title', 'description', 'image', 'icon')
        }),
        ('Moderation', {
            'fields': ('moderator',)
        }),
        ('Display Settings', {
            'fields': ('is_active', 'is_featured', 'order')
        }),
    )

    def member_count(self, obj):
        return obj.member_count
    member_count.short_description = 'Members'

# SIG Registration Admin
class SigsRegistrationAdmin(admin.ModelAdmin):
    list_display = ['user', 'sig', 'registration_date']
    list_filter = ['sig', 'registration_date']
    search_fields = ['user__email', 'user__first_name', 'user__last_name', 'sig__title']
    date_hierarchy = 'registration_date'

@admin.register(CertificateSignatory)
class CertificateSignatoryAdmin(admin.ModelAdmin):
    """Who signs the membership certificate.

    Changing the president or secretary is now an admin job: upload the new
    signature image, edit the name and title, save. Every certificate generated
    from then on carries it.
    """

    list_display = ('name', 'title', 'signature_preview', 'is_active', 'order', 'updated_at')
    list_editable = ('is_active', 'order')
    list_filter = ('is_active',)
    search_fields = ('name', 'title')
    readonly_fields = ('signature_preview_large', 'created_at', 'updated_at')

    fieldsets = (
        ('Signatory', {
            'fields': ('name', 'title'),
            'description': (
                'The name and title exactly as they should be printed on the certificate, '
                'e.g. "Dr. Kennedy Edegbe" / "National President".'
            ),
        }),
        ('Signature image', {
            'fields': ('signature', 'signature_preview_large'),
            'description': (
                'A cropped image of the signature. The certificate prints it 45pt tall and '
                'scales the width to match, so trim the whitespace around the signature.'
            ),
        }),
        ('Display', {
            'fields': ('is_active', 'order'),
        }),
        ('Record', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )

    def signature_preview(self, obj):
        if obj.signature:
            return format_html(
                '<img src="{}" style="max-height:40px; background:#fff; border:1px solid #ddd; '
                'border-radius:4px; padding:2px;"/>',
                obj.signature.url,
            )
        return format_html('<span style="color:#9ca3af;">No image</span>')
    signature_preview.short_description = 'Signature'

    def signature_preview_large(self, obj):
        if obj.signature:
            return format_html(
                '<img src="{}" style="max-height:120px; background:#fff; border:1px solid #ddd; '
                'border-radius:6px; padding:6px;"/>',
                obj.signature.url,
            )
        return format_html('<span style="color:#9ca3af;">Upload a signature image to preview it here.</span>')
    signature_preview_large.short_description = 'Preview'


admin.site.register(Sigs, SigsAdmin)
admin.site.register(SigsRegistration, SigsRegistrationAdmin)
admin.site.register(Events)
admin.site.register(News)
admin.site.register(Resource)
admin.site.register(Download)


admin.site.site_header = "ELTAN - Dashboard Admin"
admin.site.site_title ="ELTAN ADMIN"
